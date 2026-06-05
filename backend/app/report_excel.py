"""
Genera reporte Excel profesional del dashboard admin (tablas, KPIs y gráficos).
"""

from __future__ import annotations

import io
import json
from collections import defaultdict
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.legend import Legend
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import Session

from . import crud

# Paleta alineada con Pedidos Mayorista
COLOR_PRIMARY = "2ECC71"
COLOR_PRIMARY_DARK = "27AE60"
COLOR_HEADER = "1E293B"
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_KPI_BG = "ECFDF5"
COLOR_KPI_BORDER = "2ECC71"
COLOR_SUBTITLE = "64748B"
COLOR_ALT_ROW = "F8FAFC"
COLOR_META_BG = "F1F5F9"
COLOR_SECTION = "0F172A"

DASHBOARD_LAST_COL = 12
CHART_COL = 6
TABLE_STYLE = "TableStyleMedium2"


def _thin_border(color: str = "CBD5E1") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header_row(ws, row: int, col_start: int, col_end: int) -> None:
    fill = PatternFill("solid", fgColor=COLOR_HEADER)
    font = Font(name="Calibri", bold=True, color=COLOR_HEADER_TEXT, size=11)
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border(COLOR_HEADER)


def _add_excel_table(ws, ref: str, name: str) -> None:
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=TABLE_STYLE,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def _merge_range(ws, row_start: int, row_end: int, col_start: int, col_end: int) -> None:
    if row_start == row_end and col_start == col_end:
        return
    ws.merge_cells(
        start_row=row_start,
        start_column=col_start,
        end_row=row_end,
        end_column=col_end,
    )


def _ensure_aware(dt: datetime | None) -> datetime | None:
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt


def _fmt_dt(dt: datetime | None) -> str:
    if not dt:
        return "—"
    return _ensure_aware(dt).strftime("%d/%m/%Y %H:%M")


def _duration_between(start: datetime | None, end: datetime | None, now: datetime | None = None) -> str:
    if start is None:
        return "—"
    start = _ensure_aware(start)
    end = _ensure_aware(end or now or datetime.now(timezone.utc))
    total_seconds = max(0, int((end - start).total_seconds()))
    if total_seconds < 60:
        return f"{total_seconds}s"
    minutes, seconds = divmod(total_seconds, 60)
    if minutes < 60:
        return f"{minutes}m {seconds:02d}s" if seconds else f"{minutes}m"
    hours, rem_minutes = divmod(minutes, 60)
    if hours < 24:
        return f"{hours}h {rem_minutes}m" if rem_minutes else f"{hours}h"
    days = hours // 24
    return f"{days}d {hours % 24}h"


def _duration_minutes(start: datetime | None, end: datetime | None) -> float | None:
    if start is None or end is None:
        return None
    start = _ensure_aware(start)
    end = _ensure_aware(end)
    return max(0.0, (end - start).total_seconds() / 60.0)


def _pedido_estado(p) -> str:
    estado = p.estado.value if hasattr(p.estado, "value") else str(p.estado)
    return crud._ESTADO_LABELS.get(estado, estado)


def _pedido_tiempos(p, now: datetime | None = None) -> tuple[str, str, str]:
    now = now or datetime.now(timezone.utc)
    estado = p.estado.value if hasattr(p.estado, "value") else str(p.estado)
    ts, started, finished = p.timestamp, p.started_at, p.finished_at

    if estado == "pendiente":
        espera = _duration_between(ts, None, now)
        return espera, "—", espera
    if estado == "en_proceso":
        if not started:
            espera = _duration_between(ts, None, now)
            return espera, "—", espera
        return (
            _duration_between(ts, started, now),
            _duration_between(started, None, now),
            _duration_between(ts, None, now),
        )
    if estado == "finalizado":
        return (
            _duration_between(ts, started, now),
            _duration_between(started, finished, now),
            _duration_between(ts, finished, now),
        )
    return "—", "—", "—"


def _carnicero_label(user) -> str:
    if not user:
        return "—"
    name = " ".join(filter(None, [user.nombre, user.apellido])).strip()
    num = (user.numero_carnicero or user.username or "").strip()
    if num and name:
        return f"{num} — {name}"
    return name or num or "—"


def _carnicero_numero(user) -> str:
    if not user:
        return "—"
    return (user.numero_carnicero or user.username or "—").strip()


def _tiene_reporte(p) -> bool:
    if p.problema_reportado and str(p.problema_reportado).strip():
        return True
    if p.reporte_mensajes:
        try:
            msgs = json.loads(p.reporte_mensajes)
            return bool(msgs)
        except (json.JSONDecodeError, TypeError):
            pass
    return bool(p.problema_respuesta and str(p.problema_respuesta).strip())


def _reporte_resumen(p) -> str:
    partes: list[str] = []
    if p.problema_reportado and str(p.problema_reportado).strip():
        partes.append(str(p.problema_reportado).strip())
    if p.problema_respuesta and str(p.problema_respuesta).strip():
        partes.append(f"Respuesta: {p.problema_respuesta.strip()}")
    if p.reporte_mensajes:
        try:
            for msg in json.loads(p.reporte_mensajes):
                if not isinstance(msg, dict):
                    continue
                rol = msg.get("rol") or "?"
                texto = (msg.get("texto") or "").strip()
                if texto:
                    partes.append(f"[{rol}] {texto}")
        except (json.JSONDecodeError, TypeError):
            pass
    if not partes:
        return "—"
    resumen = " | ".join(partes)
    return resumen if len(resumen) <= 500 else resumen[:497] + "..."


def _write_sheet_banner(ws, title: str, last_col: int) -> None:
    letter = get_column_letter(last_col)
    ws.merge_cells(f"A1:{letter}1")
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(name="Calibri", size=14, bold=True, color=COLOR_HEADER_TEXT)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    ws.sheet_view.showGridLines = False


def _write_interactive_table(
    ws,
    *,
    title: str,
    headers: list[str],
    data_rows: list[list],
    table_name: str,
    col_widths: list[float],
    header_row: int = 3,
    note: str | None = None,
) -> None:
    last_col = len(headers)
    _write_sheet_banner(ws, title, last_col)

    if note:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        note_cell = ws.cell(row=2, column=1, value=note)
        note_cell.font = Font(size=10, color=COLOR_SUBTITLE, italic=True)
        note_cell.alignment = Alignment(wrap_text=True)

    for col, header in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=header)
    _style_header_row(ws, header_row, 1, last_col)

    rows_to_write = data_rows if data_rows else [["Sin registros para los filtros aplicados"] + [""] * (last_col - 1)]
    for i, row_values in enumerate(rows_to_write):
        r = header_row + 1 + i
        for col, value in enumerate(row_values, 1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(value, float):
                cell.number_format = "#,##0.00"

    last_row = header_row + len(rows_to_write)
    last_letter = get_column_letter(last_col)
    _add_excel_table(ws, f"A{header_row}:{last_letter}{last_row}", table_name)

    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _write_kpi_block(
    ws,
    start_row: int,
    start_col: int,
    end_col: int,
    label: str,
    value,
    unit: str = "",
) -> None:
    _merge_range(ws, start_row, start_row, start_col, end_col)
    _merge_range(ws, start_row + 1, start_row + 1, start_col, end_col)

    lbl = ws.cell(row=start_row, column=start_col, value=label)
    lbl.font = Font(name="Calibri", size=10, color=COLOR_SUBTITLE, bold=True)
    lbl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    val = ws.cell(row=start_row + 1, column=start_col, value=f"{value}{unit}")
    val.font = Font(name="Calibri", size=20, bold=True, color=COLOR_PRIMARY_DARK)
    val.alignment = Alignment(horizontal="center", vertical="center")

    fill = PatternFill("solid", fgColor=COLOR_KPI_BG)
    border = _thin_border(COLOR_KPI_BORDER)
    for r in (start_row, start_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = border


def _write_section_title(ws, row: int, col: int, text: str) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Calibri", size=12, bold=True, color=COLOR_SECTION)
    cell.border = Border(bottom=Side(style="medium", color=COLOR_PRIMARY))


def _write_data_table(
    ws,
    start_row: int,
    title: str,
    col1_header: str,
    col2_header: str,
    rows: list[dict],
    key_name: str,
    key_value: str,
    table_name: str,
) -> tuple[int, int]:
    _write_section_title(ws, start_row, 1, title)
    header_r = start_row + 1
    ws.cell(row=header_r, column=1, value=col1_header)
    ws.cell(row=header_r, column=2, value=col2_header)
    _style_header_row(ws, header_r, 1, 2)

    data_start = header_r + 1
    if not rows:
        ws.cell(row=data_start, column=1, value="Sin datos")
        ws.merge_cells(start_row=data_start, start_column=1, end_row=data_start, end_column=2)
        ws.cell(row=data_start, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=data_start, column=1).font = Font(color=COLOR_SUBTITLE, italic=True)
        return header_r, data_start

    for i, row in enumerate(rows):
        r = data_start + i
        value = row[key_value]
        if isinstance(value, float):
            value = round(value, 2)
        ws.cell(row=r, column=1, value=row[key_name])
        ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")

    data_end = data_start + len(rows) - 1
    _add_excel_table(ws, f"A{header_r}:B{data_end}", table_name)
    return header_r, data_end


def _add_bar_chart(ws, anchor_row: int, header_r: int, data_end: int, data_start: int, title: str) -> int:
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = title
    bar.y_axis.title = "Cantidad"
    bar.height = 8
    bar.width = 14
    bar.varyColors = True
    bar.legend = None

    data_ref = Reference(ws, min_col=2, min_row=header_r, max_row=data_end)
    cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, f"{get_column_letter(CHART_COL)}{anchor_row}")
    return anchor_row + 16


def _add_pie_chart(
    ws,
    anchor_row: int,
    header_r: int,
    data_end: int,
    data_start: int,
    title: str,
) -> None:
    pie = PieChart()
    pie.title = title
    pie.height = 9
    pie.width = 14
    pie.legend = Legend()
    pie.legend.position = "r"

    pie_data = Reference(ws, min_col=2, min_row=header_r, max_row=data_end)
    pie_labels = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_labels)
    ws.add_chart(pie, f"{get_column_letter(CHART_COL)}{anchor_row}")


def _build_report_rows(pedidos: list) -> dict:
    now = datetime.now(timezone.utc)
    line_items: list[list] = []
    pedido_rows: list[list] = []
    clientes: dict[str, dict] = defaultdict(
        lambda: {
            "pedidos": set(),
            "kg": 0.0,
            "productos": set(),
            "mayoristas": set(),
            "sede": "",
            "ciudad": "",
        }
    )
    carniceros: dict[str, dict] = defaultdict(
        lambda: {
            "pedidos": 0,
            "kg": 0.0,
            "prep_minutes": [],
            "sede": "",
        }
    )
    incidencias: list[list] = []

    total_lineas = 0
    pedidos_con_reporte = 0
    prep_minutes_all: list[float] = []

    for p in pedidos:
        estado = _pedido_estado(p)
        espera, prep, total = _pedido_tiempos(p, now)
        kg_total = sum(d.cantidad_kg or 0 for d in (p.detalles or []))
        tiene_reporte = _tiene_reporte(p)
        if tiene_reporte:
            pedidos_con_reporte += 1
            incidencias.append([
                p.numero_pedido or p.id,
                _fmt_dt(p.timestamp),
                p.cliente_nombre or "—",
                p.sede.nombre if p.sede else "—",
                p.sede.ciudad if p.sede and p.sede.ciudad else "—",
                estado,
                _carnicero_label(p.carnicero),
                _reporte_resumen(p),
            ])

        prep_min = _duration_minutes(p.started_at, p.finished_at)
        if prep_min is not None:
            prep_minutes_all.append(prep_min)

        carn_key = _carnicero_label(p.carnicero)
        if p.carnicero_id:
            carniceros[carn_key]["pedidos"] += 1
            carniceros[carn_key]["kg"] += kg_total
            carniceros[carn_key]["sede"] = p.sede.nombre if p.sede else "—"
            if prep_min is not None:
                carniceros[carn_key]["prep_minutes"].append(prep_min)

        cliente_key = (p.cliente_nombre or "—").strip()
        clientes[cliente_key]["pedidos"].add(p.id)
        clientes[cliente_key]["kg"] += kg_total
        clientes[cliente_key]["sede"] = p.sede.nombre if p.sede else "—"
        clientes[cliente_key]["ciudad"] = p.sede.ciudad if p.sede and p.sede.ciudad else "—"
        if p.mayorista and p.mayorista.username:
            clientes[cliente_key]["mayoristas"].add(p.mayorista.username)

        pedido_rows.append([
            p.numero_pedido or p.id,
            _fmt_dt(p.timestamp),
            p.sede.nombre if p.sede else "—",
            p.sede.ciudad if p.sede and p.sede.ciudad else "—",
            p.mayorista.username if p.mayorista else "—",
            p.cliente_nombre or "—",
            estado,
            _carnicero_label(p.carnicero),
            _carnicero_numero(p.carnicero),
            len(p.detalles or []),
            round(kg_total, 2),
            _fmt_dt(p.started_at),
            _fmt_dt(p.finished_at),
            espera,
            prep,
            total,
            (p.observaciones or "—").strip() or "—",
            "Sí" if tiene_reporte else "No",
        ])

        detalles = p.detalles or []
        if not detalles:
            line_items.append([
                p.numero_pedido or p.id,
                _fmt_dt(p.timestamp),
                p.sede.nombre if p.sede else "—",
                p.sede.ciudad if p.sede and p.sede.ciudad else "—",
                p.mayorista.username if p.mayorista else "—",
                p.cliente_nombre or "—",
                estado,
                "—",
                "—",
                "—",
                0,
                "—",
                _carnicero_label(p.carnicero),
                _carnicero_numero(p.carnicero),
                _fmt_dt(p.started_at),
                _fmt_dt(p.finished_at),
                espera,
                prep,
                total,
                (p.observaciones or "—").strip() or "—",
                "Sí" if tiene_reporte else "No",
            ])
            continue

        for d in detalles:
            total_lineas += 1
            corte = d.corte
            categoria = corte.categoria.nombre if corte and corte.categoria else "—"
            producto = corte.nombre if corte else "—"
            tipo = d.tipo_corte.nombre if d.tipo_corte else "—"
            clientes[cliente_key]["productos"].add(producto)
            line_items.append([
                p.numero_pedido or p.id,
                _fmt_dt(p.timestamp),
                p.sede.nombre if p.sede else "—",
                p.sede.ciudad if p.sede and p.sede.ciudad else "—",
                p.mayorista.username if p.mayorista else "—",
                p.cliente_nombre or "—",
                estado,
                categoria,
                producto,
                tipo,
                round(d.cantidad_kg or 0, 2),
                (d.observaciones or "—").strip() or "—",
                _carnicero_label(p.carnicero),
                _carnicero_numero(p.carnicero),
                _fmt_dt(p.started_at),
                _fmt_dt(p.finished_at),
                espera,
                prep,
                total,
                (p.observaciones or "—").strip() or "—",
                "Sí" if tiene_reporte else "No",
            ])

    cliente_rows = sorted(
        [
            [
                cliente,
                info["sede"],
                info["ciudad"],
                len(info["pedidos"]),
                round(info["kg"], 2),
                len(info["productos"]),
                ", ".join(sorted(info["mayoristas"])) or "—",
            ]
            for cliente, info in clientes.items()
        ],
        key=lambda r: (-float(r[4]), r[0]),
    )

    def _avg_minutes(values: list[float]) -> str:
        if not values:
            return "—"
        avg = sum(values) / len(values)
        if avg < 60:
            return f"{avg:.0f} min"
        hours = int(avg // 60)
        mins = int(round(avg % 60))
        return f"{hours}h {mins}m" if mins else f"{hours}h"

    carnicero_rows = sorted(
        [
            [
                nombre,
                nombre.split(" — ")[0] if " — " in nombre else "—",
                info["sede"],
                info["pedidos"],
                round(info["kg"], 2),
                round(info["kg"] / info["pedidos"], 2) if info["pedidos"] else 0,
                _avg_minutes(info["prep_minutes"]),
            ]
            for nombre, info in carniceros.items()
            if nombre != "—"
        ],
        key=lambda r: (-float(r[4]), r[0]),
    )

    avg_prep = _avg_minutes(prep_minutes_all)

    return {
        "line_items": line_items,
        "pedido_rows": pedido_rows,
        "cliente_rows": cliente_rows,
        "carnicero_rows": carnicero_rows,
        "incidencias": incidencias,
        "total_lineas": total_lineas,
        "pedidos_con_reporte": pedidos_con_reporte,
        "avg_prep": avg_prep,
    }


def build_dashboard_report(
    db: Session,
    *,
    sede_ids: list[int] | None,
    date_from,
    date_to,
    period_label: str,
    sede_label: str,
) -> tuple[bytes, str]:
    data = crud.gather_dashboard_report_data(db, sede_ids, date_from, date_to)
    detail = _build_report_rows(data["pedidos"])
    wb = Workbook()

    # --- Hoja Dashboard ---
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    last_col_letter = get_column_letter(DASHBOARD_LAST_COL)
    ws.merge_cells(f"A1:{last_col_letter}1")
    title = ws["A1"]
    title.value = "PEDIDOS MAYORISTA — Reporte operativo"
    title.font = Font(name="Calibri", size=18, bold=True, color=COLOR_HEADER_TEXT)
    title.fill = PatternFill("solid", fgColor=COLOR_PRIMARY_DARK)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    generated = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")
    meta_rows = [
        ("Generado", generated),
        ("Periodo", period_label),
        ("Sedes", sede_label),
    ]
    meta_fill = PatternFill("solid", fgColor=COLOR_META_BG)
    for i, (label, value) in enumerate(meta_rows, start=3):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=DASHBOARD_LAST_COL)
        lbl = ws.cell(row=i, column=1, value=f"{label}:")
        lbl.font = Font(bold=True, color=COLOR_SUBTITLE)
        lbl.fill = meta_fill
        ws.cell(row=i, column=2).fill = meta_fill
        val = ws.cell(row=i, column=3, value=value)
        val.font = Font(color=COLOR_SECTION)
        val.fill = meta_fill
        for c in range(4, DASHBOARD_LAST_COL + 1):
            ws.cell(row=i, column=c).fill = meta_fill

    _write_kpi_block(ws, 7, 1, 3, "PEDIDOS TOTALES", data["total_pedidos"])
    _write_kpi_block(ws, 7, 4, 6, "KG TOTAL", round(data["total_kg"], 2), " kg")
    _write_kpi_block(ws, 7, 7, 9, "KG PROMEDIO / PEDIDO", data["avg_kg"], " kg")
    _write_kpi_block(ws, 7, 10, 12, "LÍNEAS DE PRODUCTO", detail["total_lineas"])
    _write_kpi_block(ws, 10, 1, 3, "MAYORISTAS", data["mayoristas_count"])
    _write_kpi_block(ws, 10, 4, 6, "CIUDADES", data["ciudades_count"])
    _write_kpi_block(ws, 10, 7, 9, "PREP. PROMEDIO", detail["avg_prep"])
    _write_kpi_block(ws, 10, 10, 12, "PEDIDOS CON REPORTE", detail["pedidos_con_reporte"])

    content_row = 13
    use_estado = bool(data["orders_by_estado"])
    chart_rows = data["orders_by_estado"] if use_estado else data["sede_orders"]
    chart_title = "Pedidos por estado" if use_estado else "Pedidos por sede"

    summary_header, summary_end = _write_data_table(
        ws,
        content_row,
        chart_title.upper(),
        "Categoría",
        "Cantidad",
        chart_rows,
        "name",
        "count",
        "TablaResumen",
    )
    summary_data_start = summary_header + 1

    next_chart_row = content_row
    if chart_rows:
        next_chart_row = _add_bar_chart(
            ws, content_row, summary_header, summary_end, summary_data_start, chart_title
        )

    cuts_start = summary_end + 3
    cuts_header, cuts_end = _write_data_table(
        ws,
        cuts_start,
        "TOP CORTES (KG)",
        "Corte",
        "Total kg",
        data["top_cuts"],
        "name",
        "total_kg",
        "TablaCortes",
    )
    cuts_data_start = cuts_header + 1

    if data["top_cuts"]:
        pie_row = max(cuts_start, next_chart_row + 1)
        _add_pie_chart(ws, pie_row, cuts_header, cuts_end, cuts_data_start, "Distribución por corte (kg)")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    for col in range(CHART_COL, DASHBOARD_LAST_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    ws.page_setup.orientation = "landscape"

    table_note = (
        "Use los filtros en el encabezado de cada tabla (▼) para buscar por cliente, producto, "
        "carnicero, sede u otros campos."
    )

    # --- Detalle productos (línea a línea) ---
    ws_prod = wb.create_sheet("Detalle productos")
    _write_interactive_table(
        ws_prod,
        title="DETALLE DE PRODUCTOS POR CLIENTE",
        note=table_note,
        headers=[
            "Nº pedido",
            "Fecha pedido",
            "Sede",
            "Ciudad",
            "Mayorista",
            "Cliente",
            "Estado",
            "Categoría",
            "Producto",
            "Preparación",
            "Cantidad (kg)",
            "Obs. línea",
            "Carnicero",
            "Nº carnicero",
            "Inicio preparación",
            "Finalizado",
            "Tiempo espera",
            "Tiempo preparación",
            "Tiempo total",
            "Obs. pedido",
            "Tiene reporte",
        ],
        data_rows=detail["line_items"],
        table_name="TablaDetalleProductos",
        col_widths=[12, 18, 18, 14, 14, 22, 12, 14, 22, 16, 12, 24, 22, 12, 18, 18, 14, 16, 14, 24, 12],
    )

    # --- Detalle pedidos ---
    ws_ped = wb.create_sheet("Detalle pedidos")
    _write_interactive_table(
        ws_ped,
        title="DETALLE DE PEDIDOS",
        note=table_note,
        headers=[
            "Nº pedido",
            "Fecha creación",
            "Sede",
            "Ciudad",
            "Mayorista",
            "Cliente",
            "Estado",
            "Carnicero",
            "Nº carnicero",
            "Ítems",
            "Kg total",
            "Inicio preparación",
            "Finalizado",
            "Tiempo espera",
            "Tiempo preparación",
            "Tiempo total",
            "Observaciones",
            "Tiene reporte",
        ],
        data_rows=detail["pedido_rows"],
        table_name="TablaDetallePedidos",
        col_widths=[12, 18, 18, 14, 14, 22, 12, 22, 12, 8, 10, 18, 18, 14, 16, 14, 28, 12],
    )

    # --- Por cliente ---
    ws_cli = wb.create_sheet("Por cliente")
    _write_interactive_table(
        ws_cli,
        title="RESUMEN POR CLIENTE",
        note="Totales agregados por nombre de cliente en el periodo filtrado.",
        headers=[
            "Cliente",
            "Sede",
            "Ciudad",
            "Pedidos",
            "Kg total",
            "Productos distintos",
            "Mayoristas",
        ],
        data_rows=detail["cliente_rows"],
        table_name="TablaPorCliente",
        col_widths=[26, 18, 14, 10, 12, 16, 24],
    )

    # --- Por carnicero ---
    ws_carn = wb.create_sheet("Por carnicero")
    _write_interactive_table(
        ws_carn,
        title="DESPACHO POR CARNICERO",
        note="Pedidos asignados a cada carnicero con totales y tiempo promedio de preparación.",
        headers=[
            "Carnicero",
            "Nº carnicero",
            "Sede",
            "Pedidos despachados",
            "Kg total",
            "Kg promedio / pedido",
            "Prep. promedio",
        ],
        data_rows=detail["carnicero_rows"],
        table_name="TablaPorCarnicero",
        col_widths=[26, 12, 18, 16, 12, 16, 14],
    )

    # --- Incidencias ---
    ws_inc = wb.create_sheet("Incidencias")
    _write_interactive_table(
        ws_inc,
        title="REPORTES E INCIDENCIAS",
        note="Pedidos con problemas reportados o conversación de reporte activa.",
        headers=[
            "Nº pedido",
            "Fecha",
            "Cliente",
            "Sede",
            "Ciudad",
            "Estado",
            "Carnicero",
            "Resumen del reporte",
        ],
        data_rows=detail["incidencias"],
        table_name="TablaIncidencias",
        col_widths=[12, 18, 22, 18, 14, 12, 22, 48],
    )

    # --- Filtros e índice ---
    ws_filt = wb.create_sheet("Filtros")
    ws_filt.sheet_view.showGridLines = False
    ws_filt.merge_cells("A1:B1")
    ws_filt["A1"] = "FILTROS E ÍNDICE DEL REPORTE"
    ws_filt["A1"].font = Font(name="Calibri", size=14, bold=True, color=COLOR_HEADER_TEXT)
    ws_filt["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws_filt["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_filt["A3"] = "Parámetro"
    ws_filt["B3"] = "Valor"
    _style_header_row(ws_filt, 3, 1, 2)
    rows_meta = [
        ("Aplicación", "Pedidos Mayorista"),
        ("Periodo", period_label),
        ("Sedes incluidas", sede_label),
        ("Fecha desde", str(date_from) if date_from else "—"),
        ("Fecha hasta", str(date_to) if date_to else "—"),
        ("Total pedidos", data["total_pedidos"]),
        ("Kg total", round(data["total_kg"], 2)),
        ("Líneas de producto", detail["total_lineas"]),
        ("Pedidos con reporte", detail["pedidos_con_reporte"]),
        ("Preparación promedio", detail["avg_prep"]),
    ]
    for i, (k, v) in enumerate(rows_meta, 4):
        ws_filt.cell(row=i, column=1, value=k)
        ws_filt.cell(row=i, column=2, value=v)
        ws_filt.cell(row=i, column=1).font = Font(bold=True, color=COLOR_SUBTITLE)

    idx_start = 4 + len(rows_meta) + 2
    ws_filt.cell(row=idx_start, column=1, value="Hojas del reporte").font = Font(
        bold=True, size=12, color=COLOR_SECTION
    )
    hojas = [
        ("Dashboard", "Resumen ejecutivo con KPIs, gráficos y totales."),
        ("Detalle productos", "Cada producto de cada pedido: cliente, cantidad, carnicero y tiempos."),
        ("Detalle pedidos", "Una fila por pedido con tiempos, carnicero y totales."),
        ("Por cliente", "Totales agregados por cliente."),
        ("Por carnicero", "Despachos y rendimiento por carnicero."),
        ("Incidencias", "Pedidos con reportes o problemas."),
    ]
    for j, (nombre, desc) in enumerate(hojas, idx_start + 1):
        ws_filt.cell(row=j, column=1, value=nombre).font = Font(bold=True)
        ws_filt.cell(row=j, column=2, value=desc)

    ws_filt.column_dimensions["A"].width = 24
    ws_filt.column_dimensions["B"].width = 56

    buffer = io.BytesIO()
    wb.save(buffer)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"reporte_pedidos_mayorista_{stamp}.xlsx"
    return buffer.getvalue(), filename
