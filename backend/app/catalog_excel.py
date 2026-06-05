"""
Exportación e importación Excel del catálogo por sede.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import Session, joinedload

from . import crud, models, schemas

COLOR_HEADER = "1E293B"
COLOR_HEADER_TEXT = "FFFFFF"
TABLE_STYLE = "TableStyleMedium2"

SHEET_CATEGORIAS = "Categorías"
SHEET_PREPARACIONES = "Preparaciones"
SHEET_PRODUCTOS = "Productos"
SHEET_INSTRUCCIONES = "Instrucciones"


def _norm_name(value) -> str:
    return str(value or "").strip()


def _norm_key(value) -> str:
    return _norm_name(value).casefold()


def _style_header_row(ws, row: int, col_start: int, col_end: int) -> None:
    fill = PatternFill("solid", fgColor=COLOR_HEADER)
    font = Font(bold=True, color=COLOR_HEADER_TEXT)
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _add_table(ws, ref: str, name: str) -> None:
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=TABLE_STYLE,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def _write_instructions_sheet(ws) -> None:
    ws["A1"] = "Cómo usar este archivo"
    ws["A1"].font = Font(bold=True, size=14)
    lines = [
        "1. Complete las hojas Categorías, Preparaciones y Productos.",
        "2. Al importar, los elementos que ya existen (mismo nombre) no se duplican.",
        "3. En Productos, la columna Categoría debe coincidir con una categoría existente o de la hoja Categorías.",
        "4. Preparaciones: nombres separados por coma (ej: Mariposa, Delgado).",
        "5. Puede exportar el catálogo actual, editarlo y volver a cargarlo.",
    ]
    for i, line in enumerate(lines, start=3):
        ws[f"A{i}"] = line
    ws.column_dimensions["A"].width = 72


def _sheet_rows(ws, min_cols: int) -> list[tuple]:
    rows: list[tuple] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        padded = list(row) + [None] * max(0, min_cols - len(row))
        rows.append(tuple(padded[:min_cols]))
    return rows


def build_catalog_export(
    db: Session,
    sede_id: int,
    sede_nombre: str,
) -> tuple[bytes, str]:
    categorias = crud.get_categories(db, sede_id)
    tipos = crud.get_tipos_corte(db, sede_id)
    cortes = crud.get_cortes(db, sede_id)
    cat_map = {c.id: c.nombre for c in categorias}

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = SHEET_INSTRUCCIONES
    _write_instructions_sheet(ws_info)

    ws_cat = wb.create_sheet(SHEET_CATEGORIAS)
    ws_cat.append(["Nombre", "Imagen URL"])
    _style_header_row(ws_cat, 1, 1, 2)
    for cat in categorias:
        ws_cat.append([cat.nombre, cat.imagen_url or ""])
    if categorias:
        _add_table(ws_cat, f"A1:B{len(categorias) + 1}", "TablaCategorias")
    ws_cat.column_dimensions["A"].width = 28
    ws_cat.column_dimensions["B"].width = 40

    ws_prep = wb.create_sheet(SHEET_PREPARACIONES)
    ws_prep.append(["Nombre"])
    _style_header_row(ws_prep, 1, 1, 1)
    for tipo in tipos:
        ws_prep.append([tipo.nombre])
    if tipos:
        _add_table(ws_prep, f"A1:A{len(tipos) + 1}", "TablaPreparaciones")
    ws_prep.column_dimensions["A"].width = 28

    ws_prod = wb.create_sheet(SHEET_PRODUCTOS)
    ws_prod.append(["Nombre", "Categoría", "Imagen URL", "Preparaciones"])
    _style_header_row(ws_prod, 1, 1, 4)
    for corte in cortes:
        prep_names = ", ".join(t.nombre for t in (corte.tipos_corte or []))
        ws_prod.append([
            corte.nombre,
            cat_map.get(corte.categoria_id, ""),
            corte.imagen_url or "",
            prep_names,
        ])
    if cortes:
        _add_table(ws_prod, f"A1:D{len(cortes) + 1}", "TablaProductos")
    for col, width in zip("ABCD", [28, 22, 40, 32]):
        ws_prod.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(ch if ch.isalnum() else "_" for ch in sede_nombre)[:30] or "sede"
    filename = f"catalogo_{safe_name}_{stamp}.xlsx"
    return buffer.getvalue(), filename


def build_catalog_template(sede_nombre: str) -> tuple[bytes, str]:
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = SHEET_INSTRUCCIONES
    _write_instructions_sheet(ws_info)

    ws_cat = wb.create_sheet(SHEET_CATEGORIAS)
    ws_cat.append(["Nombre", "Imagen URL"])
    _style_header_row(ws_cat, 1, 1, 2)
    ws_cat.append(["Res", ""])
    _add_table(ws_cat, "A1:B2", "TablaCategorias")
    ws_cat.column_dimensions["A"].width = 28
    ws_cat.column_dimensions["B"].width = 40

    ws_prep = wb.create_sheet(SHEET_PREPARACIONES)
    ws_prep.append(["Nombre"])
    _style_header_row(ws_prep, 1, 1, 1)
    ws_prep.append(["Mariposa"])
    ws_prep.append(["Delgado"])
    _add_table(ws_prep, "A1:A3", "TablaPreparaciones")
    ws_prep.column_dimensions["A"].width = 28

    ws_prod = wb.create_sheet(SHEET_PRODUCTOS)
    ws_prod.append(["Nombre", "Categoría", "Imagen URL", "Preparaciones"])
    _style_header_row(ws_prod, 1, 1, 4)
    ws_prod.append(["Posta negra", "Res", "", "Mariposa, Delgado"])
    _add_table(ws_prod, "A1:D2", "TablaProductos")
    for col, width in zip("ABCD", [28, 22, 40, 32]):
        ws_prod.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"plantilla_catalogo_{stamp}.xlsx"
    return buffer.getvalue(), filename


def _find_category(db: Session, sede_id: int, nombre: str):
    key = _norm_key(nombre)
    if not key:
        return None
    for cat in crud.get_categories(db, sede_id):
        if _norm_key(cat.nombre) == key:
            return cat
    return None


def _find_tipo(db: Session, sede_id: int, nombre: str):
    key = _norm_key(nombre)
    if not key:
        return None
    for tipo in crud.get_tipos_corte(db, sede_id):
        if _norm_key(tipo.nombre) == key:
            return tipo
    return None


def _find_corte(db: Session, sede_id: int, categoria_id: int, nombre: str):
    key = _norm_key(nombre)
    if not key:
        return None
    for corte in crud.get_cortes(db, sede_id, categoria_id):
        if _norm_key(corte.nombre) == key:
            return corte
    return None


def _entry(tipo: str, nombre: str, motivo: str, fila: int, **extra) -> dict:
    row = {
        "tipo": tipo,
        "nombre": nombre,
        "motivo": motivo,
        "fila": fila,
    }
    row.update(extra)
    return row


def import_catalog_from_excel(db: Session, sede_id: int, content: bytes) -> dict:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Archivo Excel no válido: {exc}") from exc

    created = {"categorias": [], "tipos_corte": [], "cortes": []}
    skipped: list[dict] = []
    errors: list[dict] = []

    cat_cache: dict[str, models.Categoria] = {}
    tipo_cache: dict[str, models.TipoCorte] = {}

    for cat in crud.get_categories(db, sede_id):
        cat_cache[_norm_key(cat.nombre)] = cat
    for tipo in crud.get_tipos_corte(db, sede_id):
        tipo_cache[_norm_key(tipo.nombre)] = tipo

    # --- Preparaciones ---
    if SHEET_PREPARACIONES in wb.sheetnames:
        ws = wb[SHEET_PREPARACIONES]
        for idx, (nombre_raw,) in enumerate(_sheet_rows(ws, 1), start=2):
            nombre = _norm_name(nombre_raw)
            if not nombre:
                continue
            key = _norm_key(nombre)
            existente = tipo_cache.get(key) or _find_tipo(db, sede_id, nombre)
            if existente:
                tipo_cache[_norm_key(existente.nombre)] = existente
                skipped.append(_entry(
                    "tipo_corte", nombre, "Ya existe en el catálogo", idx,
                    existente_id=existente.id,
                ))
                continue
            try:
                nuevo = crud.create_tipo_corte(db, schemas.TipoCorteBase(nombre=nombre), sede_id)
                tipo_cache[_norm_key(nuevo.nombre)] = nuevo
                created["tipos_corte"].append({"nombre": nuevo.nombre, "id": nuevo.id})
            except Exception as exc:
                errors.append(_entry("tipo_corte", nombre, str(exc), idx))

    # --- Categorías ---
    if SHEET_CATEGORIAS in wb.sheetnames:
        ws = wb[SHEET_CATEGORIAS]
        for idx, (nombre_raw, imagen_raw) in enumerate(_sheet_rows(ws, 2), start=2):
            nombre = _norm_name(nombre_raw)
            if not nombre:
                continue
            key = _norm_key(nombre)
            imagen_url = _norm_name(imagen_raw) or None
            existente = cat_cache.get(key) or _find_category(db, sede_id, nombre)
            if existente:
                cat_cache[_norm_key(existente.nombre)] = existente
                skipped.append(_entry(
                    "categoria", nombre, "Ya existe en el catálogo", idx,
                    imagen_url=imagen_url or "",
                    existente_id=existente.id,
                ))
                continue
            try:
                nuevo = crud.create_category(
                    db,
                    schemas.CategoriaBase(nombre=nombre, imagen_url=imagen_url),
                    sede_id,
                )
                cat_cache[_norm_key(nuevo.nombre)] = nuevo
                created["categorias"].append({"nombre": nuevo.nombre, "id": nuevo.id})
            except Exception as exc:
                errors.append(_entry("categoria", nombre, str(exc), idx, imagen_url=imagen_url or ""))

    # --- Productos ---
    if SHEET_PRODUCTOS in wb.sheetnames:
        ws = wb[SHEET_PRODUCTOS]
        for idx, (nombre_raw, cat_raw, imagen_raw, prep_raw) in enumerate(_sheet_rows(ws, 4), start=2):
            nombre = _norm_name(nombre_raw)
            categoria_nombre = _norm_name(cat_raw)
            imagen_url = _norm_name(imagen_raw) or None
            preparaciones = _norm_name(prep_raw)

            if not nombre:
                continue
            if not categoria_nombre:
                errors.append(_entry(
                    "corte", nombre, "Falta el nombre de la categoría", idx,
                    categoria="", imagen_url=imagen_url or "", preparaciones=preparaciones,
                ))
                continue

            categoria = cat_cache.get(_norm_key(categoria_nombre)) or _find_category(db, sede_id, categoria_nombre)
            if not categoria:
                errors.append(_entry(
                    "corte", nombre, f"Categoría «{categoria_nombre}» no encontrada", idx,
                    categoria=categoria_nombre, imagen_url=imagen_url or "", preparaciones=preparaciones,
                ))
                continue

            existente = _find_corte(db, sede_id, categoria.id, nombre)
            if existente:
                skipped.append(_entry(
                    "corte", nombre, "Ya existe en el catálogo", idx,
                    categoria=categoria_nombre,
                    imagen_url=imagen_url or "",
                    preparaciones=preparaciones,
                    existente_id=existente.id,
                ))
                continue

            tipo_ids: list[int] = []
            prep_error = False
            if preparaciones:
                for prep_name in [p.strip() for p in preparaciones.split(",") if p.strip()]:
                    tipo = tipo_cache.get(_norm_key(prep_name)) or _find_tipo(db, sede_id, prep_name)
                    if not tipo:
                        errors.append(_entry(
                            "corte", nombre,
                            f"Preparación «{prep_name}» no encontrada", idx,
                            categoria=categoria_nombre,
                            imagen_url=imagen_url or "",
                            preparaciones=preparaciones,
                        ))
                        prep_error = True
                        break
                    tipo_ids.append(tipo.id)

            if prep_error:
                continue

            try:
                nuevo = crud.create_corte(
                    db,
                    schemas.CorteBase(
                        nombre=nombre,
                        categoria_id=categoria.id,
                        imagen_url=imagen_url,
                        tipos_corte_ids=tipo_ids,
                    ),
                    sede_id,
                )
                created["cortes"].append({
                    "nombre": nuevo.nombre,
                    "categoria": categoria_nombre,
                    "id": nuevo.id,
                })
            except Exception as exc:
                errors.append(_entry(
                    "corte", nombre, str(exc), idx,
                    categoria=categoria_nombre,
                    imagen_url=imagen_url or "",
                    preparaciones=preparaciones,
                ))

    wb.close()

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "totals": {
            "creados": (
                len(created["categorias"])
                + len(created["tipos_corte"])
                + len(created["cortes"])
            ),
            "omitidos": len(skipped),
            "errores": len(errors),
        },
    }
